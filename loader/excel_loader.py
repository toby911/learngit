import pandas as pd
import csv
from typing import Dict, List, Optional
from langchain.docstore.document import Document
from langchain.document_loaders.base import BaseLoader
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class XLSXLoader(BaseLoader):
    """Loads a XLSX file into a list of documents.

    Each document represents one row of the XLSX file. Every row is converted into a
    key/value pair and outputted to a new line in the document's page_content.

    The source for each document loaded from XLSX is set to the value of the
    `file_path` argument for all doucments by default.
    
    """

    def __init__(
        self,
        file_path: str,
        encoding: Optional[str] = None,
    ):
        self.file_path = file_path
        self.encoding = encoding

    # 获取所有sheet的名称
    def _get_sheet_names(self, excel_file):
        return excel_file.sheet_names
    
    def _read_sheet_data(self, excel_file, file_name, sheet_name):
        header, header_index = self._get_headers(excel_file, sheet_name)
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_index)
        columns = df.columns
        data = df.values.tolist()
        docs = []
        for row in data:
            row_data = {}
            for i, cell in enumerate(row):
                row_data[columns[i]] = cell
            content = str(row_data).replace('{', '').replace('}', '').replace("'", '')
            metadata = {"file_path": file_name, "sheet_name": sheet_name, "row": i}
            doc = Document(page_content=content, metadata=metadata)
            docs.append(doc)
        return docs

    def _read_sheet_head(self, excel_file, file_name, sheet_name):
        docs = []
        excel_data = pd.read_excel(excel_file, sheet_name=sheet_name)
        headers = list(excel_data.columns)
        new_lst = [str(item) if isinstance(item, int) else item for item in headers]
        
        content = "-".join(new_lst)
        metadata = {"file_path": file_name, "sheet_name": sheet_name, "row": 0}
        doc = Document(page_content=content, metadata=metadata)
        print(doc)
        docs.append(doc)
        return docs
                
    def _get_headers(self, excel_file, sheet_name):
        # 加载Excel文件
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        # 用于保存表头的列表
        headers = []
        header_index = 0
        # 读取前10行数据
        data = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, nrows=10)
        # 遍历前10行数据
        for index, row in data.iterrows():
            print(row)
            # 检查当前行是否包含None
            if not row.isnull().any():
                # 如果当前行没有None，将其作为表头
                headers = row.tolist()
                header_index = index 
                break

        # 输出提取到的表头
        result = [str(int(item)) if isinstance(item, float) else str(item) for item in headers]
        print("表头：", result, header_index)
        return result,header_index
 
 

    def _get_header2(self, excel_file, sheet_name):
        # 打开Excel文件
        workbook = load_workbook(excel_file)

        # 获取第一个工作表
        sheet = workbook.active

        # 初始化表头列表
        headers = []
        # 记录已读取的行数
        row_count = 0

        # 遍历工作表的行
        for row in sheet.iter_rows():
            # 如果已经读取了10行，则跳出循环
            if row_count >= 10:
                break

            # 检查是否为空行，如果为空行则跳过
            if all(cell.value is None for cell in row):
                continue

            for cell in row:
                # 获取合并单元格的值
                if isinstance(cell, openpyxl.cell.MergedCell) and cell.value is None:
                    value = sheet.cell(cell.row, cell.column).value
                else:
                    value = cell.value
                row_values.append(value)

            # # 处理跨行合并的情况
            # row_values = []
            # for cell in row:
            #     cell_value = cell.value
            #     for merged_range in sheet.merged_cells.ranges:
            #         if cell.coordinate in merged_range:
            #             cell_value = sheet.cell(merged_range.min_row, cell.column).value
            #             break
            #     row_values.append(cell_value)

            # 检查是否没有None值
            if all(cell.value is not None for cell in row):
                headers.append([cell.value for cell in row])
                row_count += 1

        print(headers)
        return headers       

    def load(self) -> List[Document]:
        """Load data into document objects."""
        docs = []
        excel_file = pd.ExcelFile(self.file_path)
        sheet_names = self._get_sheet_names(excel_file)
        for s in sheet_names:
            d = self._read_sheet_data(excel_file, self.file_path, s)
            docs.extend(d)
        for i in docs:
            i.metadata['page_number'] = 1
        return docs


if __name__ == "__main__":
    filename= r'/home/llp/work_temps/202304月统计(1).xlsx'
    loader = XLSXLoader(filename)
    # docs = loader.load()
    loader._get_header2(filename, '票价')
    # print(docs)

